from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from unittest.mock import patch

from django.core.exceptions import ValidationError
from django.db import IntegrityError, OperationalError
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from accounts.models import User
from courses.models import Course, Lesson, LessonStep, Subject
from curriculum_standards.models import (
    CurriculumDocumentType,
    CurriculumNodeType,
    CurriculumStandard,
    CurriculumStandardNode,
    CurriculumStandardVersion,
    CurriculumVersionStatus,
    SchoolStage,
)
from learning_analytics.evaluation_models import (
    EvaluationPlan,
    EvaluationPlanVersion,
    EvaluationReviewStatus,
    EvaluationScope,
    EvaluationCriterionVersion,
    EvaluationStandard,
    EvaluationStandardVersion,
    EvaluationTrialRecord,
)
from school.models import School


class EvaluationManagementApiTests(TestCase):
    def setUp(self):
        self.school = School.objects.create(name="Evaluation School", code="EVALUATION")
        self.other_school = School.objects.create(name="Other School", code="EVALUATION-OTHER")
        self.subject = Subject.objects.create(
            school=self.school,
            name="Information Technology",
            code="IT",
        )
        self.other_subject = Subject.objects.create(
            school=self.other_school,
            name="Other Subject",
            code="OTHER",
        )
        self.teacher = User.objects.create_user(
            username="evaluation_teacher",
            password="Teacher123!",
            role=User.Role.TEACHER,
            school=self.school,
        )
        self.other_teacher = User.objects.create_user(
            username="other_evaluation_teacher",
            password="Teacher123!",
            role=User.Role.TEACHER,
            school=self.other_school,
        )
        self.peer_teacher = User.objects.create_user(
            username="peer_evaluation_teacher",
            password="Teacher123!",
            role=User.Role.TEACHER,
            school=self.school,
        )
        self.student = User.objects.create_user(
            username="evaluation_student",
            password="123456",
            role=User.Role.STUDENT,
            school=self.school,
        )
        self.school_admin = User.objects.create_user(
            username="evaluation_admin",
            password="Admin123!",
            role=User.Role.SCHOOL_ADMIN,
            school=self.school,
        )
        self.other_school_admin = User.objects.create_user(
            username="other_evaluation_admin",
            password="Admin123!",
            role=User.Role.SCHOOL_ADMIN,
            school=self.other_school,
        )
        self.curriculum_standard = CurriculumStandard.objects.create(
            title="Information Technology Curriculum Standard",
            document_type=CurriculumDocumentType.SUBJECT_STANDARD,
            school_stage=SchoolStage.SENIOR_HIGH,
            subject_code="information_technology",
            subject_name="Information Technology",
            created_by=self.school_admin,
            updated_by=self.school_admin,
        )
        self.curriculum_version = CurriculumStandardVersion.objects.create(
            source=self.curriculum_standard,
            version_label="2025",
            publication_year=2025,
            effective_year=2025,
            title_snapshot=self.curriculum_standard.title,
            official_title="Information Technology Curriculum Standard (2025)",
            document_type_snapshot=self.curriculum_standard.document_type,
            school_stage_snapshot=self.curriculum_standard.school_stage,
            subject_code_snapshot=self.curriculum_standard.subject_code,
            subject_name_snapshot=self.curriculum_standard.subject_name,
            pdf_file="curriculum_standards/tests/information-technology-2025.pdf",
            pdf_sha256="1" * 64,
            pdf_size_bytes=1024,
            pdf_page_count=100,
            content_hash="2" * 64,
            created_by=self.school_admin,
        )
        node_definitions = (
            (CurriculumNodeType.CORE_COMPETENCY, "IT.CORE", "Core competency"),
            (CurriculumNodeType.COURSE_OBJECTIVE, "IT.OBJECTIVE", "Course objective"),
            (CurriculumNodeType.COURSE_CONTENT, "IT.CONTENT", "Course content"),
            (CurriculumNodeType.ACADEMIC_QUALITY, "IT.QUALITY", "Academic quality"),
        )
        self.curriculum_nodes = []
        for index, (node_type, code, title) in enumerate(node_definitions, start=1):
            self.curriculum_nodes.append(
                CurriculumStandardNode.objects.create(
                    version=self.curriculum_version,
                    node_type=node_type,
                    code=code,
                    title=title,
                    content=f"Published source text for {title} and data representation learning.",
                    source_page_start=index,
                    source_page_end=index,
                    source_paragraph=title,
                    sort_order=index,
                )
            )
        CurriculumStandardVersion.objects.filter(pk=self.curriculum_version.pk).update(
            status=CurriculumVersionStatus.PUBLISHED,
            reviewed_by=self.school_admin,
            published_by=self.school_admin,
        )
        CurriculumStandard.objects.filter(pk=self.curriculum_standard.pk).update(
            current_version=self.curriculum_version,
        )
        self.curriculum_version.refresh_from_db()
        self.curriculum_standard.refresh_from_db()
        self.curriculum_node_ids = [node.id for node in self.curriculum_nodes]
        self.course = Course.objects.create(
            subject=self.subject,
            title="Data and Computing",
            teacher=self.teacher,
            is_active=True,
        )
        self.other_course = Course.objects.create(
            subject=self.other_subject,
            title="Other Course",
            teacher=self.other_teacher,
            is_active=True,
        )
        self.peer_course = Course.objects.create(
            subject=self.subject,
            title="Peer Teacher Course",
            teacher=self.peer_teacher,
            is_active=True,
        )
        self.lesson = Lesson.objects.create(
            course=self.course,
            title="Evaluation lesson",
            is_active=True,
        )
        self.lesson_step = LessonStep.objects.create(
            lesson=self.lesson,
            title="Evaluation step",
            step_type=LessonStep.StepType.EVALUATION,
            created_by=self.teacher,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.teacher)

    def plan_payload(self) -> dict:
        return {
            "course": self.course.id,
            "title": "Data representation evaluation plan",
            "content_version": "2026.1",
            "target_students": "Grade 10 students studying data representation",
            "learning_goal": "Students can select representations and explain why they fit a data problem.",
            "learning_goals": [
                {
                    "code": "C1",
                    "title": "Representation selection",
                    "description": "The student selects a defensible representation for the stated data problem.",
                    "curriculum_node_ids": self.curriculum_node_ids,
                }
            ],
            "evaluation_basis": [
                {
                    "code": "E1",
                    "goal_codes": ["C1"],
                    "description": "The artifact and explanation jointly show a reasoned representation choice.",
                    "source_types": ["student artifact", "written explanation"],
                }
            ],
            "learning_activities": [
                {
                    "code": "A1",
                    "title": "Campus data inquiry",
                    "goal_codes": ["C1"],
                    "description": "Create a visualization and explain the mapping between variables and visual encodings.",
                }
            ],
            "evaluation_tasks": [
                {
                    "code": "T1",
                    "title": "Campus data visualization artifact",
                    "goal_codes": ["C1"],
                    "activity_codes": ["A1"],
                    "mode": "project",
                    "evidence_ownership": "individual",
                    "material_types": ["artifact", "observation", "oral_defense"],
                    "weight": 100,
                    "description": "Submit the visualization artifact together with a reasoned design explanation.",
                }
            ],
            "assessment_modes": ["project"],
            "content_scope": ["data representation", "visual encoding"],
            "thinking_requirements": ["apply", "analyze"],
            "support_options": ["teacher-provided data dictionary"],
            "scoring_rules": {
                "approach": "separate criteria",
                "decision_rule": "Interpret each criterion separately and do not replace missing evidence with a low score.",
            },
            "follow_up_suggestion": "Use the weakest evidenced criterion to select the next feedback prompt and practice task.",
            "curriculum_node_ids": self.curriculum_node_ids,
        }

    def standard_payload(
        self,
        plan_id: int,
        *,
        plan_version_id: int | None = None,
    ) -> dict:
        if plan_version_id is None:
            plan_version_id = EvaluationPlanVersion.objects.filter(
                source_id=plan_id
            ).latest("version_no").id
        return {
            "plan_version": plan_version_id,
            "title": "Data representation evaluation standard",
            "evaluation_target": "Student visualization artifact and written design explanation",
            "criteria": [
                {
                    "code": "D1",
                    "dimension": "subject_practice",
                    "title": "Representation reasoning",
                    "evaluation_target": "The submitted visualization and explanation",
                    "evaluation_sources": ["visualization artifact", "written explanation"],
                    "learning_goal_codes": ["C1"],
                    "evaluation_task_codes": ["T1"],
                    "evidence_ownership": "individual",
                    "material_types": ["artifact"],
                    "expected_performance": "The student links data characteristics, visual encoding choices, and the intended reader.",
                    "skip_condition": "Do not evaluate this criterion when no visualization or explanation is available.",
                    "support_options": ["data dictionary", "chart-type reference sheet"],
                    "common_problems": ["A polished chart without an explanation does not demonstrate reasoning."],
                    "level_descriptions": {
                        "1": "The representation conflicts with the data type and no defensible reason is provided.",
                        "2": "The representation is partly usable, but the explanation relies on unsupported preferences.",
                        "3": "The representation fits the main data type and the explanation identifies one relevant design reason.",
                        "4": "The representation fits the data and audience, with connected reasons for the main visual encodings.",
                        "5": "The representation is precise and the explanation evaluates alternatives, trade-offs, and audience needs.",
                    },
                    "scoring_examples": [
                        {
                            "level": 2,
                            "title": "Preference-only explanation",
                            "example_description": "The chart is readable, but the student only states that it looks better.",
                            "file_reference": "pilot-anchor-D1-L2",
                        },
                        {
                            "level": 4,
                            "title": "Connected encoding explanation",
                            "example_description": "The student connects variable types, axis choices, color, and the intended audience.",
                            "file_reference": "pilot-anchor-D1-L4",
                        },
                    ],
                    "follow_up_suggestion": "Ask the student to compare the chosen representation with one plausible alternative.",
                }
            ],
        }

    def two_task_plan_payload(self) -> dict:
        payload = self.plan_payload()
        payload["learning_goals"].append(
            {
                "code": "C2",
                "title": "Representation justification",
                "description": "The student compares alternatives and justifies a representation for its intended audience.",
                "curriculum_node_ids": self.curriculum_node_ids,
            }
        )
        payload["evaluation_basis"].append(
            {
                "code": "E2",
                "goal_codes": ["C2"],
                "description": "The explanation contains a comparison of alternatives and an audience-based justification.",
                "source_types": ["oral explanation", "design rationale"],
            }
        )
        payload["learning_activities"].append(
            {
                "code": "A2",
                "title": "Representation comparison discussion",
                "goal_codes": ["C2"],
                "description": "Compare two plausible representations and discuss their trade-offs for the intended audience.",
            }
        )
        payload["evaluation_tasks"][0]["weight"] = 50
        payload["evaluation_tasks"].append(
            {
                "code": "T2",
                "title": "Representation design defense",
                "goal_codes": ["C2"],
                "activity_codes": ["A2"],
                "mode": "project",
                "evidence_ownership": "group",
                "material_types": ["artifact", "observation", "oral_defense"],
                "weight": 50,
                "description": "The group submits its design and explains why it is appropriate for the intended audience.",
            }
        )
        return payload

    def create_plan_from_payload(self, payload: dict) -> dict:
        response = self.client.post(
            "/api/v1/teacher/evaluations/plans/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        return response.data["data"]

    def review_plan_id(self, plan_id: int):
        return self.client.post(
            f"/api/v1/teacher/evaluations/plans/{plan_id}/review-confirm/",
            {},
            format="json",
        )

    def publish_plan_id(self, plan_id: int):
        reviewed = self.review_plan_id(plan_id)
        self.assertEqual(reviewed.status_code, 200, reviewed.data)
        return self.client.post(
            f"/api/v1/teacher/evaluations/plans/{plan_id}/publish/",
            {},
            format="json",
        )

    def create_standard_from_payload(self, payload: dict) -> dict:
        response = self.client.post(
            "/api/v1/teacher/evaluations/standards/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        return response.data["data"]

    def review_standard_id(self, standard_id: int):
        return self.client.post(
            f"/api/v1/teacher/evaluations/standards/{standard_id}/review-confirm/",
            {},
            format="json",
        )

    def publish_standard_id(self, standard_id: int):
        reviewed = self.review_standard_id(standard_id)
        self.assertEqual(reviewed.status_code, 200, reviewed.data)
        return self.client.post(
            f"/api/v1/teacher/evaluations/standards/{standard_id}/publish/",
            {},
            format="json",
        )

    def create_plan(self, *, course: Course | None = None) -> dict:
        payload = self.plan_payload()
        if course is not None:
            payload["course"] = course.id
        response = self.client.post(
            "/api/v1/teacher/evaluations/plans/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        return response.data["data"]

    def test_plan_draft_api_derives_assessment_modes_from_evaluation_tasks(self):
        payload = self.plan_payload()
        payload["assessment_modes"] = ["test", "operation"]

        plan = self.create_plan_from_payload(payload)

        self.assertEqual(plan["assessment_modes"], ["project"])
        source = EvaluationPlan.objects.get(pk=plan["id"])
        self.assertEqual(source.assessment_modes, ["project"])

    def test_review_confirmation_is_teacher_scoped_audited_and_required(self):
        plan = self.create_plan()
        publish_url = f"/api/v1/teacher/evaluations/plans/{plan['id']}/publish/"

        response = self.client.post(publish_url, {}, format="json")
        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(EvaluationPlanVersion.objects.exists())

        peer_client = APIClient()
        peer_client.force_authenticate(self.peer_teacher)
        response = peer_client.post(
            f"/api/v1/teacher/evaluations/plans/{plan['id']}/review-confirm/",
            {},
            format="json",
        )
        self.assertEqual(response.status_code, 404, response.data)

        response = self.review_plan_id(plan["id"])
        self.assertEqual(response.status_code, 200, response.data)
        plan_source = EvaluationPlan.objects.get(pk=plan["id"])
        self.assertEqual(plan_source.review_status, EvaluationReviewStatus.REVIEWED)
        self.assertEqual(plan_source.reviewed_by_id, self.teacher.id)
        self.assertIsNotNone(plan_source.reviewed_at)
        self.assertEqual(len(plan_source.reviewed_content_hash), 64)

        response = self.client.post(publish_url, {}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        plan_version = EvaluationPlanVersion.objects.get(source_id=plan["id"])
        self.assertEqual(
            plan_source.reviewed_content_hash,
            plan_version.content_hash,
        )

        standard = self.create_standard_from_payload(
            self.standard_payload(plan["id"], plan_version_id=plan_version.id)
        )
        standard_publish_url = (
            f"/api/v1/teacher/evaluations/standards/{standard['id']}/publish/"
        )
        response = self.client.post(standard_publish_url, {}, format="json")
        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(EvaluationStandardVersion.objects.exists())

        response = peer_client.post(
            f"/api/v1/teacher/evaluations/standards/{standard['id']}/review-confirm/",
            {},
            format="json",
        )
        self.assertEqual(response.status_code, 404, response.data)

        response = self.review_standard_id(standard["id"])
        self.assertEqual(response.status_code, 200, response.data)
        standard_source = EvaluationStandard.objects.get(pk=standard["id"])
        self.assertEqual(
            standard_source.review_status,
            EvaluationReviewStatus.REVIEWED,
        )
        self.assertEqual(standard_source.reviewed_by_id, self.teacher.id)
        self.assertIsNotNone(standard_source.reviewed_at)
        self.assertEqual(len(standard_source.reviewed_content_hash), 64)

        response = self.client.patch(
            f"/api/v1/teacher/evaluations/standards/{standard['id']}/",
            {"title": "Revised data representation evaluation standard"},
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        standard_source.refresh_from_db()
        self.assertEqual(standard_source.review_status, EvaluationReviewStatus.DRAFT)
        self.assertIsNone(standard_source.reviewed_by_id)
        self.assertIsNone(standard_source.reviewed_at)
        self.assertEqual(standard_source.reviewed_content_hash, "")

        response = self.client.post(standard_publish_url, {}, format="json")
        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(EvaluationStandardVersion.objects.exists())

    def test_all_six_evaluation_mode_chains_can_be_published(self):
        cases = (
            ("test", ["answer", "score"], ["answer"], None),
            ("operation", ["operation", "observation"], ["operation"], None),
            (
                "project",
                ["artifact", "observation", "oral_defense"],
                ["artifact"],
                None,
            ),
            ("artifact", ["artifact", "observation"], ["artifact"], None),
            (
                "oral_defense",
                ["oral_defense", "observation"],
                ["oral_defense"],
                None,
            ),
            (
                "mixed",
                ["artifact", "observation", "oral_defense"],
                ["artifact"],
                ["artifact", "oral_defense"],
            ),
        )
        for index, (
            mode,
            task_materials,
            criterion_materials,
            component_modes,
        ) in enumerate(
            cases,
            start=1,
        ):
            with self.subTest(mode=mode):
                plan_payload = self.plan_payload()
                plan_payload["title"] = f"P2 {mode} evaluation plan {index}"
                plan_payload["assessment_modes"] = [mode]
                plan_payload["evaluation_tasks"][0].update(
                    {
                        "mode": mode,
                        "evidence_ownership": "individual",
                        "material_types": task_materials,
                    }
                )
                if component_modes is not None:
                    plan_payload["evaluation_tasks"][0][
                        "component_modes"
                    ] = component_modes
                plan = self.create_plan_from_payload(plan_payload)
                published_plan = self.publish_plan_id(plan["id"])
                self.assertEqual(published_plan.status_code, 200, published_plan.data)
                if mode == "mixed":
                    plan_version = EvaluationPlanVersion.objects.get(
                        source_id=plan["id"]
                    )
                    self.assertEqual(
                        plan_version.evaluation_tasks[0]["component_modes"],
                        component_modes,
                    )

                standard_payload = self.standard_payload(plan["id"])
                standard_payload["title"] = f"P2 {mode} evaluation standard {index}"
                standard_payload["criteria"][0].update(
                    {
                        "evidence_ownership": "individual",
                        "material_types": criterion_materials,
                    }
                )
                standard = self.create_standard_from_payload(standard_payload)
                published_standard = self.publish_standard_id(standard["id"])
                self.assertEqual(
                    published_standard.status_code,
                    200,
                    published_standard.data,
                )

    def test_mode_material_contract_and_mixed_components_are_enforced(self):
        cases = (
            {
                "mode": "test",
                "material_types": ["artifact"],
            },
            {
                "mode": "mixed",
                "material_types": ["artifact", "oral_defense"],
            },
            {
                "mode": "mixed",
                "component_modes": ["artifact"],
                "material_types": ["artifact"],
            },
            {
                "mode": "mixed",
                "component_modes": ["artifact", "mixed"],
                "material_types": ["artifact", "oral_defense"],
            },
            {
                "mode": "mixed",
                "component_modes": ["artifact", "oral_defense"],
                "material_types": ["artifact"],
            },
            {
                "mode": "project",
                "component_modes": ["artifact", "oral_defense"],
                "material_types": ["artifact", "oral_defense"],
            },
        )
        for index, task_changes in enumerate(cases, start=1):
            with self.subTest(case=index):
                payload = self.plan_payload()
                payload["title"] = f"Invalid mode contract {index}"
                payload["assessment_modes"] = [task_changes["mode"]]
                payload["evaluation_tasks"][0].update(task_changes)
                plan = self.create_plan_from_payload(payload)

                response = self.review_plan_id(plan["id"])

                self.assertEqual(response.status_code, 400, response.data)
                source = EvaluationPlan.objects.get(pk=plan["id"])
                self.assertEqual(source.review_status, EvaluationReviewStatus.DRAFT)
                self.assertIsNone(source.reviewed_by_id)
                self.assertIsNone(source.reviewed_at)
                self.assertEqual(source.reviewed_content_hash, "")
        self.assertFalse(EvaluationPlanVersion.objects.exists())

    def test_evaluation_task_and_activity_must_share_each_task_goal(self):
        cases = (
            {
                "goal_codes": ["C1"],
                "activity_codes": ["A1", "A2"],
            },
            {
                "goal_codes": ["C1", "C2"],
                "activity_codes": ["A1"],
            },
        )
        for index, task_links in enumerate(cases, start=1):
            with self.subTest(case=index):
                payload = self.two_task_plan_payload()
                payload["title"] = f"Invalid activity goal alignment {index}"
                payload["evaluation_tasks"][0].update(task_links)
                plan = self.create_plan_from_payload(payload)

                response = self.review_plan_id(plan["id"])

                self.assertEqual(response.status_code, 400, response.data)
                self.assertIn("evaluation_tasks", response.data["errors"])
        self.assertFalse(EvaluationPlanVersion.objects.exists())

    def test_standard_cannot_bind_codes_that_exist_only_in_plan_draft(self):
        plan = self.create_plan()
        published = self.publish_plan_id(plan["id"])
        self.assertEqual(published.status_code, 200, published.data)

        draft_payload = self.two_task_plan_payload()
        response = self.client.patch(
            f"/api/v1/teacher/evaluations/plans/{plan['id']}/",
            draft_payload,
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)

        standard_payload = self.standard_payload(plan["id"])
        standard_payload["criteria"][0].update(
            {
                "learning_goal_codes": ["C2"],
                "evaluation_task_codes": ["T2"],
                "evidence_ownership": "group",
                "material_types": ["artifact"],
            }
        )
        standard = self.create_standard_from_payload(standard_payload)
        rejected = self.review_standard_id(standard["id"])
        self.assertEqual(rejected.status_code, 400, rejected.data)
        self.assertIn("未知学习目标", str(rejected.data["errors"]))

    def test_standard_rejects_uncovered_evaluation_task(self):
        plan = self.create_plan_from_payload(self.two_task_plan_payload())
        published = self.publish_plan_id(plan["id"])
        self.assertEqual(published.status_code, 200, published.data)

        standard = self.create_standard_from_payload(self.standard_payload(plan["id"]))
        rejected = self.review_standard_id(standard["id"])
        self.assertEqual(rejected.status_code, 400, rejected.data)
        self.assertIn("T2", str(rejected.data["errors"]))
        self.assertIn("尚未设置评价指标", str(rejected.data["errors"]))

    def test_standard_rejects_goal_task_mismatch(self):
        plan = self.create_plan_from_payload(self.two_task_plan_payload())
        published = self.publish_plan_id(plan["id"])
        self.assertEqual(published.status_code, 200, published.data)
        payload = self.standard_payload(plan["id"])
        payload["criteria"][0].update(
            {
                "learning_goal_codes": ["C1"],
                "evaluation_task_codes": ["T2"],
                "evidence_ownership": "group",
                "material_types": ["artifact"],
            }
        )
        standard = self.create_standard_from_payload(payload)
        rejected = self.review_standard_id(standard["id"])
        self.assertEqual(rejected.status_code, 400, rejected.data)
        self.assertIn("必须来自其所选评价任务", str(rejected.data["errors"]))

    def test_standard_rejects_ownership_and_material_type_mismatches(self):
        plan = self.create_plan()
        published = self.publish_plan_id(plan["id"])
        self.assertEqual(published.status_code, 200, published.data)

        ownership_payload = self.standard_payload(plan["id"])
        ownership_payload["title"] = "Ownership mismatch standard"
        ownership_payload["criteria"][0]["evidence_ownership"] = "group"
        ownership_standard = self.create_standard_from_payload(ownership_payload)
        ownership_rejected = self.review_standard_id(ownership_standard["id"])
        self.assertEqual(ownership_rejected.status_code, 400, ownership_rejected.data)
        self.assertIn("材料归属", str(ownership_rejected.data["errors"]))

        material_payload = self.standard_payload(plan["id"])
        material_payload["title"] = "Material mismatch standard"
        material_payload["criteria"][0]["material_types"] = ["answer"]
        material_standard = self.create_standard_from_payload(material_payload)
        material_rejected = self.review_standard_id(material_standard["id"])
        self.assertEqual(material_rejected.status_code, 400, material_rejected.data)
        self.assertIn("共同的评价材料类型", str(material_rejected.data["errors"]))

    def test_both_task_requires_aggregate_individual_and_group_criterion_coverage(self):
        plan_payload = self.plan_payload()
        plan_payload["evaluation_tasks"][0]["evidence_ownership"] = "both"
        plan = self.create_plan_from_payload(plan_payload)
        published = self.publish_plan_id(plan["id"])
        self.assertEqual(published.status_code, 200, published.data)

        incomplete_payload = self.standard_payload(plan["id"])
        incomplete_payload["title"] = "Incomplete both ownership standard"
        incomplete = self.create_standard_from_payload(incomplete_payload)
        response = self.review_standard_id(incomplete["id"])
        self.assertEqual(response.status_code, 400, response.data)
        self.assertIn("criteria", response.data["errors"])

        aggregate_payload = self.standard_payload(plan["id"])
        aggregate_payload["title"] = "Aggregate both ownership standard"
        group_criterion = deepcopy(aggregate_payload["criteria"][0])
        group_criterion.update(
            {
                "code": "D2",
                "title": "Group representation reasoning",
                "evaluation_target": "The group visualization and shared rationale",
                "evidence_ownership": "group",
            }
        )
        aggregate_payload["criteria"].append(group_criterion)
        aggregate = self.create_standard_from_payload(aggregate_payload)
        response = self.publish_standard_id(aggregate["id"])
        self.assertEqual(response.status_code, 200, response.data)

        shared_payload = self.standard_payload(plan["id"])
        shared_payload["title"] = "Single both ownership standard"
        shared_payload["criteria"][0]["evidence_ownership"] = "both"
        shared = self.create_standard_from_payload(shared_payload)
        response = self.publish_standard_id(shared["id"])
        self.assertEqual(response.status_code, 200, response.data)

    def test_standard_stays_bound_to_the_exact_selected_plan_version(self):
        plan = self.create_plan()
        response = self.publish_plan_id(plan["id"])
        self.assertEqual(response.status_code, 200, response.data)
        first_plan_version = EvaluationPlanVersion.objects.get(
            source_id=plan["id"],
            version_no=1,
        )

        revised_payload = self.plan_payload()
        revised_payload["follow_up_suggestion"] = (
            "Use a second contrast case before asking the student to revise the explanation."
        )
        response = self.client.patch(
            f"/api/v1/teacher/evaluations/plans/{plan['id']}/",
            revised_payload,
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        response = self.publish_plan_id(plan["id"])
        self.assertEqual(response.status_code, 200, response.data)
        second_plan_version = EvaluationPlanVersion.objects.get(
            source_id=plan["id"],
            version_no=2,
        )

        standard_payload = self.standard_payload(
            plan["id"],
            plan_version_id=first_plan_version.id,
        )
        standard = self.create_standard_from_payload(standard_payload)
        standard_source = EvaluationStandard.objects.get(pk=standard["id"])
        self.assertEqual(standard_source.plan_id, plan["id"])
        self.assertEqual(standard_source.plan_version_id, first_plan_version.id)

        response = self.publish_standard_id(standard["id"])
        self.assertEqual(response.status_code, 200, response.data)
        standard_version = EvaluationStandardVersion.objects.get(
            source_id=standard["id"]
        )
        self.assertEqual(standard_version.plan_version_id, first_plan_version.id)
        self.assertNotEqual(standard_version.plan_version_id, second_plan_version.id)

        response = self.client.patch(
            f"/api/v1/teacher/evaluations/standards/{standard['id']}/",
            {"plan_version": second_plan_version.id},
            format="json",
        )
        self.assertEqual(response.status_code, 400, response.data)
        standard_source.refresh_from_db()
        self.assertEqual(standard_source.plan_version_id, first_plan_version.id)

        legacy_payload = self.standard_payload(
            plan["id"],
            plan_version_id=second_plan_version.id,
        )
        legacy_payload["plan"] = plan["id"]
        response = self.client.post(
            "/api/v1/teacher/evaluations/standards/",
            legacy_payload,
            format="json",
        )
        self.assertEqual(response.status_code, 400, response.data)
        self.assertIn("unknown_fields", response.data["errors"])

    def test_standard_detail_api_returns_complete_exact_plan_version(self):
        plan = self.create_plan()
        response = self.publish_plan_id(plan["id"])
        self.assertEqual(response.status_code, 200, response.data)
        first_plan_version = EvaluationPlanVersion.objects.get(
            source_id=plan["id"],
            version_no=1,
        )

        revised_payload = self.plan_payload()
        revised_payload["learning_goals"][0]["title"] = (
            "Revised representation selection"
        )
        revised_payload["evaluation_tasks"][0]["description"] = (
            "Submit a revised visualization and compare it with one alternative."
        )
        response = self.client.patch(
            f"/api/v1/teacher/evaluations/plans/{plan['id']}/",
            revised_payload,
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        response = self.publish_plan_id(plan["id"])
        self.assertEqual(response.status_code, 200, response.data)
        second_plan_version = EvaluationPlanVersion.objects.get(
            source_id=plan["id"],
            version_no=2,
        )

        standard = self.create_standard_from_payload(
            self.standard_payload(
                plan["id"],
                plan_version_id=first_plan_version.id,
            )
        )
        response = self.client.get(
            f"/api/v1/teacher/evaluations/standards/{standard['id']}/"
        )

        self.assertEqual(response.status_code, 200, response.data)
        plan_version = response.data["data"]["plan_version"]
        self.assertEqual(
            plan_version,
            {
                "id": first_plan_version.id,
                "source_plan_id": plan["id"],
                "title": first_plan_version.title,
                "version_no": first_plan_version.version_no,
                "content_hash": first_plan_version.content_hash,
                "review_status": EvaluationReviewStatus.REVIEWED,
                "subject": {
                    "id": self.subject.id,
                    "name": self.subject.name,
                },
                "course": {
                    "id": self.course.id,
                    "title": self.course.title,
                },
                "learning_goals": first_plan_version.learning_goals,
                "evaluation_tasks": first_plan_version.evaluation_tasks,
            },
        )
        self.assertNotEqual(plan_version["id"], second_plan_version.id)

    def create_published_standard(
        self, *, course: Course | None = None
    ) -> EvaluationStandardVersion:
        plan = self.create_plan(course=course)
        response = self.publish_plan_id(plan["id"])
        self.assertEqual(response.status_code, 200, response.data)
        response = self.client.post(
            "/api/v1/teacher/evaluations/standards/",
            self.standard_payload(plan["id"]),
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        standard_id = response.data["data"]["id"]
        response = self.publish_standard_id(standard_id)
        self.assertEqual(response.status_code, 200, response.data)
        return EvaluationStandardVersion.objects.get(source_id=standard_id)

    def test_teacher_can_bind_published_standard_to_own_lesson_step(self):
        version = self.create_published_standard()
        url = f"/api/v1/teacher/evaluations/lesson-steps/{self.lesson_step.id}/binding/"

        response = self.client.patch(
            url,
            {
                "standard_version": version.id,
                "enable_self": "true",
                "enable_peer": "false",
                "enable_teacher": "true",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        binding = response.data["data"]
        self.assertTrue(binding["enable_self"])
        self.assertFalse(binding["enable_peer"])
        self.assertTrue(binding["enable_teacher"])
        self.assertFalse(binding["locked"])
        self.assertEqual(binding["criteria"][0]["code"], "D1")
        self.assertEqual(len(binding["criteria"][0]["level_descriptions"]), 5)

        loaded = self.client.get(url)
        self.assertEqual(loaded.status_code, 200, loaded.data)
        self.assertEqual(loaded.data["data"]["binding"]["standard_version"], version.id)
        use_boundaries = {
            row["code"]: row for row in loaded.data["data"]["use_boundaries"]
        }
        self.assertEqual(use_boundaries["classroom_feedback"]["status"], "available")
        self.assertEqual(
            use_boundaries["learning_state_update"]["status"], "requires_review"
        )
        self.assertEqual(
            use_boundaries["research_and_model"]["status"], "not_direct"
        )

    def test_lesson_step_binding_requires_teacher_scope_and_one_type(self):
        version = self.create_published_standard()
        url = f"/api/v1/teacher/evaluations/lesson-steps/{self.lesson_step.id}/binding/"

        response = self.client.patch(
            url,
            {
                "standard_version": version.id,
                "enable_self": "false",
                "enable_peer": "false",
                "enable_teacher": "false",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400, response.data)

        other_client = APIClient()
        other_client.force_authenticate(self.other_teacher)
        self.assertEqual(other_client.get(url).status_code, 404)

        admin_client = APIClient()
        admin_client.force_authenticate(self.school_admin)
        self.assertEqual(admin_client.get(url).status_code, 403)

    def test_lesson_step_cannot_bind_standard_from_another_course(self):
        other_course = Course.objects.create(
            subject=self.subject,
            title="Second teacher course",
            teacher=self.teacher,
            is_active=True,
        )
        other_version = self.create_published_standard(course=other_course)
        response = self.client.patch(
            f"/api/v1/teacher/evaluations/lesson-steps/{self.lesson_step.id}/binding/",
            {
                "standard_version": other_version.id,
                "enable_self": True,
                "enable_peer": False,
                "enable_teacher": True,
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400, response.data)

    def test_options_and_drafts_are_teacher_course_scoped(self):
        response = self.client.get("/api/v1/teacher/evaluations/options/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual([row["id"] for row in response.data["data"]["courses"]], [self.course.id])
        enabled = [row for row in response.data["data"]["scopes"] if row["enabled"]]
        self.assertEqual([row["value"] for row in enabled], [EvaluationScope.COURSE])

        payload = self.plan_payload()
        payload["course"] = self.other_course.id
        response = self.client.post(
            "/api/v1/teacher/evaluations/plans/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(EvaluationPlan.objects.exists())

        student_client = APIClient()
        student_client.force_authenticate(self.student)
        self.assertEqual(
            student_client.get("/api/v1/teacher/evaluations/plans/").status_code,
            403,
        )
        school_admin_client = APIClient()
        school_admin_client.force_authenticate(self.school_admin)
        self.assertEqual(
            school_admin_client.get("/api/v1/teacher/evaluations/plans/").status_code,
            403,
        )

        peer_payload = self.plan_payload()
        peer_payload["course"] = self.peer_course.id
        response = self.client.post(
            "/api/v1/teacher/evaluations/plans/",
            peer_payload,
            format="json",
        )
        self.assertEqual(response.status_code, 400)

    def test_options_api_returns_complete_p2_plan_version_contract(self):
        plan = self.create_plan()
        response = self.publish_plan_id(plan["id"])
        self.assertEqual(response.status_code, 200, response.data)
        version = EvaluationPlanVersion.objects.get(source_id=plan["id"])

        response = self.client.get("/api/v1/teacher/evaluations/options/")

        self.assertEqual(response.status_code, 200, response.data)
        course_option = next(
            row
            for row in response.data["data"]["courses"]
            if row["id"] == self.course.id
        )
        self.assertEqual(
            course_option["subject"],
            {
                "id": self.subject.id,
                "name": self.subject.name,
                "code": self.subject.code,
            },
        )
        plan_version = next(
            row
            for row in response.data["data"]["plan_versions"]
            if row["id"] == version.id
        )
        self.assertEqual(
            plan_version,
            {
                "id": version.id,
                "source_plan_id": plan["id"],
                "title": version.title,
                "version_no": version.version_no,
                "content_hash": version.content_hash,
                "review_status": EvaluationReviewStatus.REVIEWED,
                "subject": {
                    "id": self.subject.id,
                    "name": self.subject.name,
                },
                "course": {
                    "id": self.course.id,
                    "title": self.course.title,
                },
                "learning_goals": version.learning_goals,
                "evaluation_tasks": version.evaluation_tasks,
            },
        )

    def test_incomplete_draft_can_save_but_cannot_publish(self):
        response = self.client.post(
            "/api/v1/teacher/evaluations/plans/",
            {"course": self.course.id, "title": "Working draft"},
            format="json",
        )
        self.assertEqual(response.status_code, 201)
        plan_id = response.data["data"]["id"]

        response = self.client.post(
            f"/api/v1/teacher/evaluations/plans/{plan_id}/publish/",
            {},
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(EvaluationPlanVersion.objects.exists())

    def test_plan_publish_is_immutable_idempotent_and_versioned(self):
        plan = self.create_plan()
        publish_url = f"/api/v1/teacher/evaluations/plans/{plan['id']}/publish/"

        reviewed = self.review_plan_id(plan["id"])
        self.assertEqual(reviewed.status_code, 200, reviewed.data)
        response = self.client.post(publish_url, {}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        version = EvaluationPlanVersion.objects.get()
        self.assertEqual(version.version_no, 1)
        self.assertEqual(len(version.content_hash), 64)
        self.assertEqual(version.review_status, "reviewed")
        self.assertEqual(version.reviewed_by_id, self.teacher.id)
        self.assertIsNotNone(version.reviewed_at)
        self.assertEqual(version.reviewed_content_hash, version.content_hash)

        response = self.client.post(publish_url, {}, format="json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(EvaluationPlanVersion.objects.count(), 1)

        version.title = "Changed"
        with self.assertRaises(ValidationError):
            version.save()

        payload = self.plan_payload()
        payload["follow_up_suggestion"] = "Provide a contrast case and ask the student to revise the explanation before the next task."
        response = self.client.patch(
            f"/api/v1/teacher/evaluations/plans/{plan['id']}/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        plan_source = EvaluationPlan.objects.get(pk=plan["id"])
        self.assertEqual(plan_source.review_status, "draft")
        self.assertIsNone(plan_source.reviewed_by_id)
        self.assertIsNone(plan_source.reviewed_at)
        self.assertEqual(plan_source.reviewed_content_hash, "")

        response = self.client.post(publish_url, {}, format="json")
        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(EvaluationPlanVersion.objects.count(), 1)

        reviewed = self.review_plan_id(plan["id"])
        self.assertEqual(reviewed.status_code, 200, reviewed.data)
        response = self.client.post(publish_url, {}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(
            list(EvaluationPlanVersion.objects.order_by("version_no").values_list("version_no", flat=True)),
            [1, 2],
        )

    def test_plan_publish_recovers_from_one_sqlite_create_error(self):
        errors = (
            IntegrityError("forced concurrent version conflict"),
            OperationalError("database is locked"),
        )
        for index, injected_error in enumerate(errors, start=1):
            with self.subTest(error=type(injected_error).__name__):
                payload = self.plan_payload()
                payload["title"] = f"Retryable plan publish {index}"
                plan = self.create_plan_from_payload(payload)
                reviewed = self.review_plan_id(plan["id"])
                self.assertEqual(reviewed.status_code, 200, reviewed.data)

                original_create = EvaluationPlanVersion.objects.create
                create_calls = 0

                def flaky_create(*args, **kwargs):
                    nonlocal create_calls
                    create_calls += 1
                    if create_calls == 1:
                        raise injected_error
                    return original_create(*args, **kwargs)

                with patch(
                    "learning_analytics.services.evaluation."
                    "EvaluationPlanVersion.objects.create",
                    side_effect=flaky_create,
                ):
                    response = self.client.post(
                        f"/api/v1/teacher/evaluations/plans/{plan['id']}/publish/",
                        {},
                        format="json",
                    )

                self.assertEqual(response.status_code, 200, response.data)
                self.assertEqual(create_calls, 2)
                self.assertEqual(
                    EvaluationPlanVersion.objects.filter(source_id=plan["id"]).count(),
                    1,
                )

    def test_standard_publish_recovers_from_one_sqlite_create_error(self):
        plan = self.create_plan()
        response = self.publish_plan_id(plan["id"])
        self.assertEqual(response.status_code, 200, response.data)
        errors = (
            IntegrityError("forced concurrent version conflict"),
            OperationalError("database is locked"),
        )
        for index, injected_error in enumerate(errors, start=1):
            with self.subTest(error=type(injected_error).__name__):
                payload = self.standard_payload(plan["id"])
                payload["title"] = f"Retryable standard publish {index}"
                standard = self.create_standard_from_payload(payload)
                reviewed = self.review_standard_id(standard["id"])
                self.assertEqual(reviewed.status_code, 200, reviewed.data)

                original_create = EvaluationStandardVersion.objects.create
                create_calls = 0

                def flaky_create(*args, **kwargs):
                    nonlocal create_calls
                    create_calls += 1
                    if create_calls == 1:
                        raise injected_error
                    return original_create(*args, **kwargs)

                with patch(
                    "learning_analytics.services.evaluation."
                    "EvaluationStandardVersion.objects.create",
                    side_effect=flaky_create,
                ):
                    response = self.client.post(
                        f"/api/v1/teacher/evaluations/standards/{standard['id']}/publish/",
                        {},
                        format="json",
                    )

                self.assertEqual(response.status_code, 200, response.data)
                self.assertEqual(create_calls, 2)
                self.assertEqual(
                    EvaluationStandardVersion.objects.filter(
                        source_id=standard["id"]
                    ).count(),
                    1,
                )

    def test_standard_publish_creates_normalized_immutable_level_descriptions(self):
        plan = self.create_plan()
        published_plan = self.publish_plan_id(plan["id"])
        self.assertEqual(published_plan.status_code, 200, published_plan.data)
        response = self.client.post(
            "/api/v1/teacher/evaluations/standards/",
            self.standard_payload(plan["id"]),
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        standard_id = response.data["data"]["id"]

        response = self.publish_standard_id(standard_id)
        self.assertEqual(response.status_code, 200, response.data)
        version = EvaluationStandardVersion.objects.get()
        criterion = EvaluationCriterionVersion.objects.get()
        self.assertEqual(version.plan_version.version_no, 1)
        self.assertEqual(version.review_status, "reviewed")
        self.assertEqual(version.reviewed_by_id, self.teacher.id)
        self.assertIsNotNone(version.reviewed_at)
        self.assertEqual(version.reviewed_content_hash, version.content_hash)
        self.assertEqual(criterion.dimension, "subject_practice")
        self.assertEqual(criterion.scoring_examples.count(), 2)
        self.assertIn("Do not evaluate", criterion.skip_condition)

        criterion.title = "Changed"
        with self.assertRaises(ValidationError):
            criterion.save()

    def test_forbidden_operational_indicator_cannot_enter_published_standard(self):
        plan = self.create_plan()
        published_plan = self.publish_plan_id(plan["id"])
        self.assertEqual(published_plan.status_code, 200, published_plan.data)
        payload = self.standard_payload(plan["id"])
        payload["criteria"][0]["title"] = "签到与出勤表现"
        response = self.client.post(
            "/api/v1/teacher/evaluations/standards/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 201)
        standard_id = response.data["data"]["id"]
        response = self.review_standard_id(standard_id)
        self.assertEqual(response.status_code, 400)
        self.assertFalse(EvaluationStandardVersion.objects.exists())

    def test_other_teacher_cannot_read_or_publish_plans(self):
        plan = self.create_plan()
        other_client = APIClient()
        other_client.force_authenticate(self.other_teacher)
        detail_url = f"/api/v1/teacher/evaluations/plans/{plan['id']}/"
        publish_url = f"{detail_url}publish/"
        self.assertEqual(other_client.get(detail_url).status_code, 404)
        self.assertEqual(other_client.post(publish_url, {}, format="json").status_code, 404)

    def test_scope_cannot_be_submitted_directly(self):
        payload = self.plan_payload()
        payload["scope"] = EvaluationScope.ANALYSIS
        response = self.client.post(
            "/api/v1/teacher/evaluations/plans/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("unknown_fields", response.data["errors"])

    def test_trial_record_flow_and_completed_history_protection(self):
        version = self.create_published_standard()
        options = self.client.get("/api/v1/teacher/evaluations/options/")
        self.assertEqual(options.status_code, 200)
        self.assertEqual(
            [row["id"] for row in options.data["data"]["standard_versions"]],
            [version.id],
        )

        payload = {
            "standard_version": version.id,
            "record_type": "classroom_trial",
            "title": "高一数据表达课堂试用",
            "status": "planned",
            "activity_date": "2026-07-20",
            "participant_count": 0,
            "agreement_rate": None,
            "conclusion": "pending",
            "summary": "",
            "issues": [],
            "action_items": [],
        }
        response = self.client.post(
            "/api/v1/teacher/evaluations/trials/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        record_id = response.data["data"]["id"]

        payload.update(
            {
                "status": "completed",
                "participant_count": 32,
                "conclusion": "revise",
                "summary": "学生能够理解主要指标，但二星和三星说明仍需区分。",
                "issues": ["二星和三星说明接近"],
                "action_items": ["修改两档说明后重新发布标准版本"],
            }
        )
        response = self.client.patch(
            f"/api/v1/teacher/evaluations/trials/{record_id}/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["data"]["status_label"], "已完成")

        response = self.client.patch(
            f"/api/v1/teacher/evaluations/trials/{record_id}/",
            {"title": "不应修改"},
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        response = self.client.delete(
            f"/api/v1/teacher/evaluations/trials/{record_id}/"
        )
        self.assertEqual(response.status_code, 409)
        self.assertTrue(EvaluationTrialRecord.objects.filter(pk=record_id).exists())

        completed = EvaluationTrialRecord.objects.get(pk=record_id)
        self.assertEqual(len(completed.completion_hash), 64)
        self.assertEqual(completed.completed_by_id, self.teacher.id)
        self.assertIsNotNone(completed.completed_at)
        self.assertEqual(
            completed.completion_hash,
            completed.compute_completion_hash(),
        )

        completed.title = "ORM mutation must be rejected"
        with self.assertRaises(ValidationError):
            completed.save()
        with self.assertRaises(ValidationError):
            completed.delete()
        persisted = EvaluationTrialRecord.objects.get(pk=record_id)
        self.assertEqual(persisted.title, payload["title"])
        self.assertEqual(persisted.completion_hash, completed.completion_hash)

    def test_scoring_check_requires_agreement_rate(self):
        version = self.create_published_standard()
        payload = {
            "standard_version": version.id,
            "record_type": "scoring_check",
            "title": "两名教师评分检查",
            "status": "completed",
            "activity_date": "2026-07-20",
            "participant_count": 2,
            "agreement_rate": None,
            "conclusion": "ready",
            "summary": "两名教师完成同一批作品评分。",
            "issues": [],
            "action_items": [],
        }
        response = self.client.post(
            "/api/v1/teacher/evaluations/trials/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("agreement_rate", response.data["errors"])

        payload["agreement_rate"] = "87.50"
        response = self.client.post(
            "/api/v1/teacher/evaluations/trials/",
            payload,
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)

    def test_trial_records_are_teacher_scoped_and_exportable(self):
        version = self.create_published_standard()
        record = EvaluationTrialRecord.objects.create(
            school=self.school,
            standard_version=version,
            record_type="content_review",
            title="内容审核",
            status="completed",
            activity_date="2026-07-20",
            participant_count=3,
            conclusion="ready",
            summary="审核完成。",
            issues=[],
            action_items=[],
            created_by=self.teacher,
            updated_by=self.teacher,
        )
        other_client = APIClient()
        other_client.force_authenticate(self.other_teacher)
        self.assertEqual(
            other_client.get(
                f"/api/v1/teacher/evaluations/trials/{record.id}/"
            ).status_code,
            404,
        )

        response = self.client.get(
            "/api/v1/teacher/evaluations/trials/export/"
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertIn("评价试用记录", workbook.sheetnames)
        rows = list(workbook["评价试用记录"].iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "记录ID")
        self.assertEqual(rows[1][6], "内容审核")
