from __future__ import annotations

import uuid
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from accounts.models import User
from courses.models import Course, CourseClass, Subject
from learning_analytics.models import (
    AnalyticsPipelineRun,
    DataQualityReport,
    DecisionPoint,
    LongitudinalAnalysisRun,
    ModelComparisonRun,
    ModelEvaluationResult,
    LearningEventV2,
    LearningOpportunity,
    LearningOpportunityTransitionFact,
    OutcomeDefinition,
    OutcomeObservation,
    StudentFeatureSnapshot,
    TrainingDatasetVersion,
)
from learning_analytics.services.feature_registry import (
    sync_feature_and_outcome_definitions,
)
from learning_analytics.services.feature_snapshots import create_decision_point
from learning_analytics.services.outcomes import (
    build_training_dataset,
    mature_due_outcomes,
)
from learning_analytics.services.longitudinal import build_longitudinal_analysis
from learning_analytics.services.model_comparison import build_model_comparison
from learning_analytics.services.schema_registry import sync_event_schema_definitions
from school.models import ClassGroup, School, StudentProfile, TeachingAssignment


class FeatureOutcomeDatasetTests(TestCase):
    def setUp(self):
        sync_event_schema_definitions()
        sync_feature_and_outcome_definitions()
        self.school = School.objects.create(name="特征结果测试学校", code="FEATURE")
        self.class_group = ClassGroup.objects.create(
            school=self.school,
            name="高一1班",
            grade="高一",
        )
        self.subject = Subject.objects.create(
            school=self.school,
            name="信息科技",
            code="IT",
        )
        self.school_admin = User.objects.create_user(
            username="feature_admin",
            password="Admin123!",
            role=User.Role.SCHOOL_ADMIN,
            school=self.school,
        )
        self.teacher = User.objects.create_user(
            username="feature_teacher",
            password="Teacher123!",
            role=User.Role.TEACHER,
            school=self.school,
        )
        TeachingAssignment.objects.create(
            school=self.school,
            class_group=self.class_group,
            teacher=self.teacher,
        )
        self.student = self._student("feature_student")
        self.no_future_student = self._student("feature_no_future")
        self.course = Course.objects.create(
            subject=self.subject,
            title="数据与计算",
            teacher=self.teacher,
            is_active=True,
        )
        CourseClass.objects.create(
            course=self.course,
            class_group=self.class_group,
            created_by=self.teacher,
        )
        self.client = APIClient()
        self.t0 = timezone.now() - timedelta(days=8, minutes=10)
        self._quality_report(self.t0 - timedelta(days=7), self.t0)
        self._quality_report(self.t0, self.t0 + timedelta(days=7))

    def _student(self, username):
        user = User.objects.create_user(
            username=username,
            password="123456",
            role=User.Role.STUDENT,
            school=self.school,
        )
        StudentProfile.objects.create(
            user=user,
            class_group=self.class_group,
            is_first_use=False,
        )
        return user

    def _quality_report(self, window_start, window_end):
        run = AnalyticsPipelineRun.objects.create(
            school=self.school,
            pipeline_type=AnalyticsPipelineRun.PipelineType.DATA_QUALITY,
            trigger=AnalyticsPipelineRun.Trigger.MANUAL,
            status=AnalyticsPipelineRun.Status.SUCCEEDED,
            window_start=window_start,
            window_end=window_end,
            check_version="test-v1",
            config_hash="a" * 64,
            started_at=timezone.now(),
            finished_at=timezone.now(),
        )
        return DataQualityReport.objects.create(
            school=self.school,
            pipeline_run=run,
            window_start=window_start,
            window_end=window_end,
            check_version="test-v1",
            source_checksum=uuid.uuid4().hex * 2,
            status=DataQualityReport.Status.GREEN,
            checks_passed=True,
            event_count=10,
            receive_attempt_count=10,
            learning_task_link_rate=1,
            thresholds={"test": True},
            counts={"test": 10},
            issues=[],
        )

    def _post_events(self, actor, events):
        self.client.force_authenticate(actor)
        response = self.client.post(
            "/api/v1/learning-events/batch/",
            {"batch_id": str(uuid.uuid4()), "events": events},
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(
            all(
                item["status"] == "accepted"
                for item in response.data["data"]["results"]
            ),
            response.data,
        )

    def _release_question(
        self, *, student, object_id, occurred_at, due_at, backdate=True
    ):
        event_id = uuid.uuid4()
        object_version = f"{object_id}@1"
        self._post_events(
            self.teacher,
            [
                {
                    "event_id": str(event_id),
                    "event_name": "content.released",
                    "schema_version": "1.1",
                    "source": "teacher-web",
                    "class_id": self.class_group.id,
                    "subject_id": self.subject.id,
                    "course_id": self.course.id,
                    "object_type": "question",
                    "object_id": object_id,
                    "object_version": object_version,
                    "client_occurred_at": occurred_at.isoformat(),
                    "payload": {
                        "content_type": "question",
                        "required": True,
                        "available_from": occurred_at.isoformat(),
                        "available_to": due_at.isoformat(),
                        "target_layers": ["all"],
                        "target_student_ids": [student.id],
                    },
                }
            ],
        )
        if backdate:
            LearningEventV2.objects.filter(event_id=event_id).update(
                server_received_at=occurred_at + timedelta(seconds=5),
                quality_errors=[],
            )
        return LearningOpportunity.objects.get(
            student=student,
            object_id=object_id,
        )

    def _submit(self, *, student, opportunity, occurred_at):
        event_id = uuid.uuid4()
        attempt_id = uuid.uuid4()
        self._post_events(
            student,
            [
                {
                    "event_id": str(event_id),
                    "event_name": "item.submitted",
                    "schema_version": "1.0",
                    "source": "student-web",
                    "class_id": self.class_group.id,
                    "subject_id": self.subject.id,
                    "course_id": self.course.id,
                    "object_type": "question",
                    "object_id": opportunity.object_id,
                    "object_version": opportunity.object_version,
                    "opportunity_id": str(opportunity.opportunity_id),
                    "attempt_id": str(attempt_id),
                    "client_occurred_at": occurred_at.isoformat(),
                    "payload": {
                        "question_version": opportunity.object_version,
                        "response_kind": "single",
                        "attempt_no": 1,
                        "response_time_ms": 20_000,
                    },
                }
            ],
        )
        LearningEventV2.objects.filter(event_id=event_id).update(
            server_received_at=occurred_at + timedelta(seconds=5),
            quality_errors=[],
        )
        LearningOpportunityTransitionFact.objects.filter(
            source_event__event_id=event_id
        ).update(recorded_at=occurred_at + timedelta(seconds=5))

    def _build_decision_point(self):
        for index in range(3):
            occurred_at = self.t0 - timedelta(days=5 - index)
            opportunity = self._release_question(
                student=self.student,
                object_id=f"prior-{index}",
                occurred_at=occurred_at,
                due_at=self.t0 - timedelta(days=2 - index),
            )
            self._submit(
                student=self.student,
                opportunity=opportunity,
                occurred_at=opportunity.available_to - timedelta(minutes=5),
            )
        self._release_question(
            student=self.no_future_student,
            object_id="late-received-before-t0",
            occurred_at=self.t0 - timedelta(days=3),
            due_at=self.t0 - timedelta(days=1),
            backdate=False,
        )
        result = create_decision_point(
            school=self.school,
            class_group=self.class_group,
            subject=self.subject,
            course=self.course,
            scheduled_for=self.t0,
            created_by=self.school_admin,
            purpose=DecisionPoint.Purpose.PILOT,
        )
        return result["decision_point"]

    def test_snapshot_uses_only_data_known_at_t0_and_preserves_missing_codes(self):
        point = self._build_decision_point()
        snapshot = StudentFeatureSnapshot.objects.get(
            decision_point=point,
            student=self.student,
            view_type=StudentFeatureSnapshot.ViewType.OPERATIONAL,
        )
        self.assertEqual(snapshot.values["prior_due_required_count__7d"], 3)
        self.assertEqual(snapshot.denominators["opp_completion_rate__7d"], 3)
        self.assertEqual(snapshot.values["opp_completion_rate__7d"], 1.0)

        late_snapshot = StudentFeatureSnapshot.objects.get(
            decision_point=point,
            student=self.no_future_student,
            view_type=StudentFeatureSnapshot.ViewType.OPERATIONAL,
        )
        self.assertIsNone(late_snapshot.values["prior_due_required_count__7d"])
        self.assertEqual(
            late_snapshot.missing_codes["prior_due_required_count__7d"],
            "NO_OPPORTUNITY",
        )
        self.assertEqual(
            late_snapshot.source_watermark["server_received_through"],
            self.t0.isoformat(),
        )

    def test_future_unobserved_is_not_zero_and_dataset_is_repeatable(self):
        point = self._build_decision_point()
        future_opportunities = []
        for index in range(3):
            occurred_at = self.t0 + timedelta(hours=2 + index)
            opportunity = self._release_question(
                student=self.student,
                object_id=f"future-{index}",
                occurred_at=occurred_at,
                due_at=self.t0 + timedelta(days=index + 1),
            )
            future_opportunities.append(opportunity)
        self._submit(
            student=self.student,
            opportunity=future_opportunities[0],
            occurred_at=future_opportunities[0].available_to - timedelta(minutes=5),
        )
        self._submit(
            student=self.student,
            opportunity=future_opportunities[1],
            occurred_at=future_opportunities[1].available_to - timedelta(minutes=5),
        )

        counts = mature_due_outcomes(
            school=self.school,
            as_of=self.t0 + timedelta(days=9),
        )
        self.assertEqual(counts["observed"], 2)
        self.assertEqual(counts["unobserved"], 2)
        completion = OutcomeDefinition.objects.get(
            outcome_key="required_completion_next_7d",
            version="1.0",
        )
        observed = OutcomeObservation.objects.get(
            decision_point=point,
            student=self.student,
            outcome_definition=completion,
            status=OutcomeObservation.Status.OBSERVED,
        )
        self.assertEqual(observed.numerator, Decimal("2"))
        self.assertEqual(observed.denominator, Decimal("3"))
        self.assertEqual(observed.value, Decimal("0.666667"))
        unobserved = OutcomeObservation.objects.get(
            decision_point=point,
            student=self.no_future_student,
            outcome_definition=completion,
            status=OutcomeObservation.Status.UNOBSERVED,
        )
        self.assertIsNone(unobserved.value)
        self.assertEqual(unobserved.missing_code, "NO_OPPORTUNITY")

        dataset = build_training_dataset(
            school=self.school,
            subject=self.subject,
            outcome_definition=completion,
            created_by=self.school_admin,
        )
        rebuilt = build_training_dataset(
            school=self.school,
            subject=self.subject,
            outcome_definition=completion,
            created_by=self.school_admin,
        )
        self.assertEqual(dataset.id, rebuilt.id)
        self.assertEqual(dataset.row_count, 2)
        self.assertEqual(dataset.observed_count, 1)
        self.assertEqual(dataset.unobserved_count, 1)
        self.assertFalse(dataset.manifest["comparison_ready"])
        self.assertIn(
            "event_quality_flag_rate__7d",
            dataset.manifest["audit_only_feature_keys"],
        )
        self.assertNotIn(
            "event_quality_flag_rate__7d",
            dataset.manifest["model_input_feature_keys"],
        )
        self.assertIn(
            "opp_completion_rate__7d",
            dataset.manifest["model_input_feature_keys"],
        )
        self.assertNotIn(
            "opp_completion_rate",
            dataset.manifest["model_input_feature_keys"],
        )
        self.assertEqual(TrainingDatasetVersion.objects.count(), 1)

    def test_school_admin_api_is_scoped_and_export_is_pseudonymous(self):
        point = self._build_decision_point()
        mature_due_outcomes(
            school=self.school,
            as_of=self.t0 + timedelta(days=9),
        )
        completion = OutcomeDefinition.objects.get(
            outcome_key="required_completion_next_7d",
            version="1.0",
        )
        dataset = build_training_dataset(
            school=self.school,
            subject=self.subject,
            outcome_definition=completion,
            created_by=self.school_admin,
        )
        self.client.force_authenticate(self.school_admin)
        response = self.client.get("/api/v1/school-admin/analytics/preparation/")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["data"]["summary"]["decision_point_count"], 1)
        export = self.client.get(
            f"/api/v1/school-admin/analytics/preparation/datasets/{dataset.id}/export/"
        )
        self.assertEqual(export.status_code, 200)
        workbook = load_workbook(BytesIO(export.content), read_only=True)
        rows = list(workbook["匿名数据"].iter_rows(values_only=True))
        flattened = " ".join(str(cell) for row in rows for cell in row if cell)
        self.assertNotIn(self.student.username, flattened)
        self.assertNotIn(self.no_future_student.username, flattened)

        self.client.force_authenticate(self.teacher)
        self.assertEqual(
            self.client.get("/api/v1/school-admin/analytics/preparation/").status_code,
            403,
        )
        self.client.force_authenticate(self.student)
        self.assertEqual(
            self.client.get("/api/v1/school-admin/analytics/preparation/").status_code,
            403,
        )
        self.assertEqual(point.school_id, self.school.id)

    def test_longitudinal_and_model_comparison_are_repeatable_and_block_small_data(self):
        self._build_decision_point()
        future_opportunities = []
        for index in range(3):
            occurred_at = self.t0 + timedelta(hours=2 + index)
            opportunity = self._release_question(
                student=self.student,
                object_id=f"validation-future-{index}",
                occurred_at=occurred_at,
                due_at=self.t0 + timedelta(days=index + 1),
            )
            future_opportunities.append(opportunity)
        for opportunity in future_opportunities[:2]:
            self._submit(
                student=self.student,
                opportunity=opportunity,
                occurred_at=opportunity.available_to - timedelta(minutes=5),
            )
        mature_due_outcomes(school=self.school, as_of=self.t0 + timedelta(days=9))
        completion = OutcomeDefinition.objects.get(
            outcome_key="required_completion_next_7d",
            version="1.0",
        )
        dataset = build_training_dataset(
            school=self.school,
            subject=self.subject,
            outcome_definition=completion,
            created_by=self.school_admin,
        )

        longitudinal = build_longitudinal_analysis(
            dataset=dataset,
            created_by=self.school_admin,
        )
        longitudinal_again = build_longitudinal_analysis(
            dataset=dataset,
            created_by=self.school_admin,
        )
        self.assertEqual(longitudinal.id, longitudinal_again.id)
        self.assertEqual(longitudinal.status, LongitudinalAnalysisRun.Status.COMPLETED)
        self.assertTrue(
            LongitudinalAnalysisRun.objects.get(pk=longitudinal.id)
            .feature_results.exists()
        )

        comparison = build_model_comparison(
            dataset=dataset,
            created_by=self.school_admin,
        )
        comparison_again = build_model_comparison(
            dataset=dataset,
            created_by=self.school_admin,
        )
        self.assertEqual(comparison.id, comparison_again.id)
        self.assertEqual(comparison.status, ModelComparisonRun.Status.BLOCKED)
        self.assertTrue(comparison.model_card["prohibited_use"])
        self.assertEqual(comparison.negative_controls.count(), 5)
        for evaluation in comparison.evaluations.all():
            self.assertNotEqual(evaluation.status, ModelEvaluationResult.Status.READY)
            self.assertIsNone(evaluation.primary_metric)
            self.assertIsNone(evaluation.rmse)
            self.assertIsNone(evaluation.mae)
            self.assertFalse(evaluation.metrics["reportable"])

        self.client.force_authenticate(self.school_admin)
        response = self.client.get("/api/v1/school-admin/analytics/models/")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(len(response.data["data"]["comparison_runs"]), 1)
