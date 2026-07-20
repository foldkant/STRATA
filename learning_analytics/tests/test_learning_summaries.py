from __future__ import annotations

import uuid
from datetime import timedelta
from io import BytesIO

from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from accounts.models import User
from courses.models import Course, CourseClass, Subject
from learning.models import StratificationDecision
from learning_analytics.models import StudentLearningSummary
from learning_analytics.services.learning_summaries import (
    build_student_learning_summary,
    build_transparent_suggestion,
)
from learning_analytics.services.schema_registry import sync_event_schema_definitions
from school.models import ClassGroup, School, StudentProfile, TeachingAssignment


class LearningSummaryTests(TestCase):
    def setUp(self):
        sync_event_schema_definitions()
        self.school = School.objects.create(name="学习汇总测试学校", code="SUMMARY")
        self.class_group = ClassGroup.objects.create(
            school=self.school, name="高一1班", grade="高一"
        )
        self.subject = Subject.objects.create(
            school=self.school, name="信息科技", code="IT"
        )
        self.teacher = User.objects.create_user(
            username="summary_teacher",
            password="Teacher123!",
            role=User.Role.TEACHER,
            school=self.school,
        )
        self.other_teacher = User.objects.create_user(
            username="summary_other_teacher",
            password="Teacher123!",
            role=User.Role.TEACHER,
            school=self.school,
        )
        TeachingAssignment.objects.create(
            school=self.school,
            class_group=self.class_group,
            teacher=self.teacher,
        )
        self.student = User.objects.create_user(
            username="summary_student",
            password="123456",
            role=User.Role.STUDENT,
            school=self.school,
        )
        self.profile = StudentProfile.objects.create(
            user=self.student,
            class_group=self.class_group,
            current_layer="B",
            is_first_use=False,
        )
        self.course = Course.objects.create(
            subject=self.subject,
            title="数据与计算",
            teacher=self.teacher,
            is_active=True,
        )
        CourseClass.objects.create(
            course=self.course,
            class_group=self.class_group,
            created_by=self.teacher,
        )
        self.client = APIClient()
        self.base_time = timezone.now() - timedelta(hours=1)

    def post_events(self, actor, events):
        self.client.force_authenticate(actor)
        response = self.client.post(
            "/api/v1/learning-events/batch/",
            {"batch_id": str(uuid.uuid4()), "events": events},
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(
            all(item["status"] == "accepted" for item in response.data["data"]["results"]),
            response.data,
        )
        return response

    def create_scored_opportunities(self):
        for index in range(5):
            occurred_at = self.base_time + timedelta(minutes=index * 5)
            object_id = f"summary-question-{index + 1}"
            object_version = f"{object_id}@1"
            self.post_events(
                self.teacher,
                [
                    {
                        "event_id": str(uuid.uuid4()),
                        "event_name": "content.released",
                        "schema_version": "1.0",
                        "source": "teacher-web",
                        "class_id": self.class_group.id,
                        "subject_id": self.subject.id,
                        "course_id": self.course.id,
                        "object_type": "question",
                        "object_id": object_id,
                        "object_version": object_version,
                        "client_occurred_at": occurred_at.isoformat(),
                        "payload": {
                            "content_type": "question",
                            "required": True,
                            "target_layers": ["all"],
                        },
                    }
                ],
            )
            from learning_analytics.models import LearningOpportunity

            opportunity = LearningOpportunity.objects.get(
                student=self.student, object_id=object_id
            )
            attempt_id = uuid.uuid4()
            self.post_events(
                self.student,
                [
                    {
                        "event_id": str(uuid.uuid4()),
                        "event_name": "item.submitted",
                        "schema_version": "1.0",
                        "source": "student-web",
                        "class_id": self.class_group.id,
                        "subject_id": self.subject.id,
                        "course_id": self.course.id,
                        "object_type": "question",
                        "object_id": object_id,
                        "object_version": object_version,
                        "opportunity_id": str(opportunity.opportunity_id),
                        "attempt_id": str(attempt_id),
                        "client_occurred_at": (occurred_at + timedelta(minutes=1)).isoformat(),
                        "payload": {
                            "question_version": object_version,
                            "response_kind": "single",
                            "attempt_no": 1,
                            "response_time_ms": 20_000,
                        },
                    }
                ],
            )
            self.post_events(
                self.teacher,
                [
                    {
                        "event_id": str(uuid.uuid4()),
                        "event_name": "item.graded",
                        "schema_version": "1.0",
                        "source": "teacher-web",
                        "target_student_id": self.student.id,
                        "class_id": self.class_group.id,
                        "subject_id": self.subject.id,
                        "course_id": self.course.id,
                        "object_type": "question",
                        "object_id": object_id,
                        "object_version": object_version,
                        "opportunity_id": str(opportunity.opportunity_id),
                        "attempt_id": str(attempt_id),
                        "client_occurred_at": (occurred_at + timedelta(minutes=2)).isoformat(),
                        "payload": {
                            "grading_state": "final",
                            "score_raw": 4,
                            "score_max": 5,
                            "grader_type": "teacher",
                        },
                    }
                ],
            )

    def test_windows_are_repeatable_and_suggestion_stays_teacher_only(self):
        self.create_scored_opportunities()
        summaries = {}
        for window_type, _label in StudentLearningSummary.WindowType.choices:
            summaries[window_type] = build_student_learning_summary(
                student_profile=self.profile,
                course=self.course,
                window_type=window_type,
                as_of=timezone.localdate(),
            )
        summary = summaries[StudentLearningSummary.WindowType.DAYS_30]
        self.assertEqual(summary.data_status, StudentLearningSummary.DataStatus.AVAILABLE)
        self.assertEqual(summary.metrics["opportunities"]["eligible_count"], 5)
        self.assertEqual(summary.metrics["opportunities"]["submitted_count"], 5)
        self.assertEqual(summary.metrics["completion_rate"], 1.0)
        self.assertEqual(summary.metrics["score"]["score_rate"], 0.8)
        suggestion = build_transparent_suggestion(summary=summary)
        self.assertEqual(suggestion.suggested_layer, "A")
        self.assertEqual(suggestion.previous_layer, "B")

        rebuilt = build_student_learning_summary(
            student_profile=self.profile,
            course=self.course,
            window_type=StudentLearningSummary.WindowType.DAYS_30,
            as_of=timezone.localdate(),
        )
        self.assertEqual(rebuilt.id, summary.id)
        self.assertEqual(StudentLearningSummary.objects.count(), 4)
        self.assertEqual(StratificationDecision.objects.count(), 1)

        self.client.force_authenticate(self.other_teacher)
        self.assertEqual(
            self.client.get("/api/v1/teacher/analytics/stratification/").data["data"],
            [],
        )
        self.client.force_authenticate(self.student)
        self.assertEqual(
            self.client.get("/api/v1/teacher/analytics/stratification/").status_code,
            403,
        )

        self.client.force_authenticate(self.teacher)
        listed = self.client.get("/api/v1/teacher/analytics/stratification/")
        self.assertEqual(listed.status_code, 200, listed.data)
        self.assertEqual(len(listed.data["data"]), 1)
        exported = self.client.get(
            "/api/v1/teacher/analytics/learning-summaries/export/?window=30d"
        )
        self.assertEqual(exported.status_code, 200)
        self.assertEqual(
            exported["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(exported.content), read_only=True)
        sheet = workbook["学习情况"]
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.cell(2, 1).value, self.student.username)
        self.assertEqual(sheet.cell(2, 14).value, "100.0%")
        reviewed = self.client.post(
            f"/api/v1/teacher/analytics/stratification/{suggestion.id}/review/",
            {"action": "accept", "note": "先按建议安排下一阶段任务。"},
            format="json",
        )
        self.assertEqual(reviewed.status_code, 200, reviewed.data)
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.current_layer, "B")
        suggestion.refresh_from_db()
        self.assertEqual(suggestion.status, StratificationDecision.Status.ACCEPTED)
        self.assertEqual(suggestion.teacher_selected_layer, "A")

        self.client.force_authenticate(self.student)
        denied_export = self.client.get(
            "/api/v1/teacher/analytics/learning-summaries/export/?window=30d"
        )
        self.assertEqual(denied_export.status_code, 403)

    def test_no_task_is_not_counted_as_zero_score(self):
        summary = build_student_learning_summary(
            student_profile=self.profile,
            course=self.course,
            window_type=StudentLearningSummary.WindowType.DAYS_7,
            as_of=timezone.localdate(),
        )
        self.assertEqual(
            summary.data_status, StudentLearningSummary.DataStatus.NO_OPPORTUNITY
        )
        self.assertIsNone(summary.metrics["completion_rate"])
        self.assertIsNone(summary.metrics["score"]["score_rate"])
        suggestion = build_transparent_suggestion(summary=summary)
        self.assertEqual(suggestion.suggested_layer, "")
        self.assertEqual(suggestion.confidence, 0)

    def test_model_candidate_prevents_parallel_transparent_pending_suggestion(self):
        summary = build_student_learning_summary(
            student_profile=self.profile,
            course=self.course,
            window_type=StudentLearningSummary.WindowType.DAYS_30,
            as_of=timezone.localdate(),
        )
        model_candidate = StratificationDecision.objects.create(
            student=self.student,
            class_group=self.class_group,
            subject=self.subject,
            course=self.course,
            previous_layer="B",
            suggested_layer="B",
            confidence=0.7,
            window_start=summary.window_start,
            window_end=summary.window_end,
            rule_version="m03-test-candidate",
        )

        suggestion = build_transparent_suggestion(summary=summary)

        self.assertEqual(suggestion.id, model_candidate.id)
        self.assertEqual(StratificationDecision.objects.count(), 1)
