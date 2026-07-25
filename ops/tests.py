from __future__ import annotations

import json
import tempfile
import zipfile
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from accounts.models import User
from ops.models import ImportBatch
from school.models import ClassGroup, School, StudentProfile


def collection_zip(*, school_code: str, unsafe: bool = False) -> bytes:
    payload = BytesIO()
    with zipfile.ZipFile(payload, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "manifest.json",
            json.dumps(
                {
                    "school_code": school_code,
                    "system_version": "2.0-test",
                    "schema_version": "1",
                    "exported_at": "2026-07-21T08:00:00+08:00",
                }
            ),
        )
        archive.writestr("../unsafe.txt" if unsafe else "data/summary.json", "{}")
    return payload.getvalue()


@override_settings(MEDIA_ROOT=tempfile.mkdtemp(prefix="strata-ops-tests-"))
class SuperAdminConsoleApiTests(TestCase):
    def setUp(self):
        self.formal_school = School.objects.create(
            name="正式学校", code="FORMAL-01", is_synthetic=False
        )
        self.synthetic_school = School.objects.create(
            name="测试学校", code="SYNTHETIC-01", is_synthetic=True
        )
        self.super_admin = User.objects.create_superuser(
            username="super_console",
            password="Admin123!",
            role=User.Role.SUPER_ADMIN,
        )
        self.school_admin = User.objects.create_user(
            username="school_console",
            password="Admin123!",
            role=User.Role.SCHOOL_ADMIN,
            school=self.formal_school,
        )
        self.client = APIClient()

    def upload_package(self, content: bytes, name: str = "collection.zip"):
        return self.client.post(
            "/api/v1/super-admin/collection/",
            {
                "package_file": SimpleUploadedFile(
                    name, content, content_type="application/zip"
                )
            },
            format="multipart",
        )

    def test_collection_upload_validates_and_rejects_duplicate(self):
        self.client.force_authenticate(self.super_admin)
        content = collection_zip(school_code=self.formal_school.code)
        response = self.upload_package(content)

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["data"]["status"], ImportBatch.Status.VALIDATED)
        self.assertEqual(
            response.data["data"]["source_school"]["id"], self.formal_school.id
        )
        self.assertEqual(response.data["data"]["validation"]["errors"], [])

        duplicate = self.upload_package(content, "duplicate.zip")
        self.assertEqual(duplicate.status_code, 409)
        self.assertEqual(ImportBatch.objects.count(), 1)
        batch_id = response.data["data"]["id"]
        detail = self.client.get(f"/api/v1/super-admin/collection/{batch_id}/")
        exported = self.client.get("/api/v1/super-admin/collection/export/")
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(
            exported["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        deleted = self.client.delete(f"/api/v1/super-admin/collection/{batch_id}/")
        self.assertEqual(deleted.status_code, 200)
        self.assertFalse(ImportBatch.objects.exists())

    def test_collection_rejects_unsafe_archive_path(self):
        self.client.force_authenticate(self.super_admin)
        response = self.upload_package(
            collection_zip(school_code=self.formal_school.code, unsafe=True)
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["data"]["status"], ImportBatch.Status.FAILED)
        self.assertTrue(response.data["data"]["validation"]["errors"])

    def test_cross_school_scope_and_health_are_super_admin_only(self):
        self.client.force_authenticate(self.super_admin)
        formal = self.client.get("/api/v1/super-admin/analysis/")
        with_test = self.client.get(
            "/api/v1/super-admin/analysis/?include_test_data=1"
        )
        health = self.client.get("/api/v1/super-admin/health/")
        analysis_export = self.client.get("/api/v1/super-admin/analysis/export/")
        health_export = self.client.get("/api/v1/super-admin/health/export/")

        self.assertEqual(formal.status_code, 200)
        self.assertEqual(len(formal.data["data"]["schools"]), 1)
        self.assertEqual(len(with_test.data["data"]["schools"]), 2)
        self.assertNotIn("layers", formal.data["data"]["charts"])
        self.assertNotIn("school_layer_coverage", formal.data["data"]["charts"])
        self.assertNotIn("layer_coverage", formal.data["data"]["schools"][0])
        self.assertEqual(health.status_code, 200)
        self.assertTrue(health.data["data"]["checks"])
        self.assertEqual(analysis_export.status_code, 200)
        self.assertEqual(health_export.status_code, 200)
        analysis_book = load_workbook(
            BytesIO(analysis_export.content),
            read_only=True,
        )
        self.assertNotIn("当前分层", analysis_book.sheetnames)
        school_headers = [
            cell.value for cell in next(analysis_book["学校对比"].iter_rows())
        ]
        self.assertNotIn("分层覆盖率", school_headers)

        self.client.force_authenticate(self.school_admin)
        for url in (
            "/api/v1/super-admin/collection/",
            "/api/v1/super-admin/analysis/",
            "/api/v1/super-admin/health/",
        ):
            self.assertEqual(self.client.get(url).status_code, 403)

    def test_school_dashboard_uses_clear_operational_language(self):
        self.client.force_authenticate(self.school_admin)
        response = self.client.get("/api/v1/school-admin/dashboard/")

        self.assertEqual(response.status_code, 200)
        metric_labels = [item["label"] for item in response.data["data"]["metrics"]]
        reminder_labels = [
            item["label"] for item in response.data["data"]["status_rows"]
        ]
        self.assertNotIn("待处理", metric_labels)
        self.assertNotIn("停用账号", reminder_labels)
        self.assertIn("近 7 天活跃学生", metric_labels)
        self.assertNotIn("student_layers", response.data["data"]["charts"])

    def test_global_student_management_ignores_legacy_layer_filter(self):
        class_group = ClassGroup.objects.create(
            school=self.formal_school,
            name="高一1班",
            grade="高一",
        )
        for suffix, legacy_layer in (("a", "A"), ("c", "C")):
            student = User.objects.create_user(
                username=f"global_student_{suffix}",
                password="123456",
                role=User.Role.STUDENT,
                school=self.formal_school,
            )
            StudentProfile.objects.create(
                user=student,
                class_group=class_group,
                current_layer=legacy_layer,
            )

        self.client.force_authenticate(self.school_admin)
        response = self.client.get("/api/v1/school-admin/students/?layer=A")

        self.assertEqual(response.status_code, 200, response.data)
        rows = response.data["data"]["results"]
        self.assertEqual(len(rows), 2)
        self.assertTrue(all("current_layer" not in row for row in rows))
        self.assertTrue(all("current_layer_label" not in row for row in rows))
        export = self.client.get(
            "/api/v1/school-admin/students/export/?layer=A"
        )
        self.assertEqual(export.status_code, 200)
        export_book = load_workbook(BytesIO(export.content), read_only=True)
        export_rows = list(export_book.active.iter_rows(values_only=True))
        self.assertNotIn("层级", export_rows[0])
        self.assertEqual(len(export_rows), 3)

    def test_legacy_student_import_layer_is_ignored_with_warning(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "登录账号",
                "姓名",
                "学号",
                "班级",
                "联系电话",
                "初始密码",
                "层级",
                "小组号",
                "积分",
                "状态",
            ]
        )
        sheet.append(
            ["legacy_student", "测试学生", "", "", "", "123456", "A", "", "0", "启用"]
        )
        content = BytesIO()
        workbook.save(content)
        content.seek(0)

        self.client.force_authenticate(self.school_admin)
        response = self.client.post(
            "/api/v1/school-admin/students/import/",
            {
                "file": SimpleUploadedFile(
                    "legacy-students.xlsx",
                    content.getvalue(),
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data["data"]["warnings"])
        profile = StudentProfile.objects.get(user__username="legacy_student")
        self.assertIsNone(profile.current_layer)

        template = self.client.get("/api/v1/school-admin/students/template/")
        self.assertEqual(template.status_code, 200)
        template_book = load_workbook(BytesIO(template.content), read_only=True)
        values = {
            value
            for row in template_book.active.iter_rows(values_only=True)
            for value in row
            if value is not None
        }
        self.assertNotIn("层级", values)
