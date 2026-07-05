from __future__ import annotations

import json
from datetime import date, datetime
from io import BytesIO
from urllib.parse import quote

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="EAF1FF")
HEADER_FONT = Font(bold=True, color="111827")
NOTE_FILL = PatternFill("solid", fgColor="FFF4DD")
THIN_BORDER = Border(bottom=Side(style="thin", color="D7E0EC"))


def normalize_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def display_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return value
    return str(value)


def append_rows_sheet(workbook: Workbook, title: str, headers: list[str], rows) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
    for row in rows:
        sheet.append([display_value(value) for value in row])
    sheet.freeze_panes = "A2"
    _fit_columns(sheet)


def build_workbook(sheets: list[dict]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet in sheets:
        append_rows_sheet(workbook, sheet["title"], sheet["headers"], sheet["rows"])
    return workbook


def workbook_response(workbook: Workbook, filename: str) -> HttpResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response


def export_rows(filename: str, title: str, headers: list[str], rows) -> HttpResponse:
    workbook = build_workbook([{"title": title, "headers": headers, "rows": rows}])
    return workbook_response(workbook, filename)


def template_response(
    filename: str,
    title: str,
    headers: list[str],
    examples: list[list[str]],
    *,
    instructions: list[str] | None = None,
    dropdowns: dict[str, list[str]] | None = None,
) -> HttpResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入模板"

    instructions = instructions or []
    if instructions:
        sheet.append(["填写说明"])
        sheet["A1"].fill = NOTE_FILL
        sheet["A1"].font = HEADER_FONT
        for text in instructions:
            sheet.append([text])
        start_row = len(instructions) + 3
    else:
        start_row = 1

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")

    for example in examples:
        sheet.append(example)

    sheet.freeze_panes = f"A{start_row + 1}"
    sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{max(start_row + len(examples), start_row)}"

    for header, values in (dropdowns or {}).items():
        if header not in headers:
            continue
        column = get_column_letter(headers.index(header) + 1)
        formula = '"' + ",".join(values) + '"'
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{column}{start_row + 1}:{column}500")

    _fit_columns(sheet)
    return workbook_response(workbook, filename)


def read_table_rows(uploaded_file, *, required_headers: list[str], all_headers: list[str] | None = None) -> list[dict]:
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
    sheet = workbook.active
    header_row_number = 0
    headers = []
    for row_number, raw_headers in enumerate(sheet.iter_rows(values_only=True), start=1):
        possible_headers = [normalize_text(value) for value in raw_headers]
        if all(header in possible_headers for header in required_headers):
            headers = possible_headers
            header_row_number = row_number
            break
    if not headers:
        raise ValueError("Excel 文件缺少表头：" + "、".join(required_headers))

    known_headers = all_headers or headers
    results = []
    for index, raw_row in enumerate(sheet.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
        values = [normalize_text(value) for value in raw_row]
        if not any(values):
            continue
        row = {"__row_number": index}
        for header in known_headers:
            row[header] = ""
        for header, value in zip(headers, values):
            if header in row:
                row[header] = value
        results.append(row)
    return results


def _fit_columns(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 10
        for cell in column_cells:
            value = normalize_text(cell.value)
            if value:
                max_length = max(max_length, min(len(value) + 2, 42))
        sheet.column_dimensions[column_letter].width = max_length
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
